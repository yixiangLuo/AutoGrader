#!/usr/bin/python
# -*- coding: utf-8 -*-
# Author: Yixiang 2019/09/28

from copy import deepcopy
import os.path
import sys
import re
import glob
import math
import string
from openpyxl import load_workbook

def convert_int(s):
    try:
        number = math.ceil(s)
        return number
    except:
        return "no"

def form_alias(l):
    return [[item] for item in l]

def exit_alert(msg):
    print(msg)
    sys.exit()

def insert_list_1D(t_list, t_index, insert_data):
    if t_list[t_index] != 0:
        print(t_list[t_index], insert_data)
        exit_alert("find multiple keywords in the same table!")
    else: t_list[t_index] = insert_data

# imported from previous project China_Eco
def data_table(file_path, keywords_1s, keywords_2, scores, anchor_1=[[]], anchor_2=[[]]):
    # MODIFIED: add a parameter "scores", modify the structure of keywords_1 to be keywords_1s, so as to include each worksheet
    ''' keywords_1 and keywords_2 are 2D list, with the outer list representing keywords,
    and the inner list representing the alias of a certain keyword.
    anchor refers to additional constraint for a keyword. It is a 2D list.
    the outer list element refers to each constraint, and the inner list takes
    the form of [int(relative coodinate in row axis), int(relative coodinate in col axis),
    str(value in the anchor cell)]'''
    # open xlsx file and prepare for data container
    wb = load_workbook(filename=file_path)
    # MODIFIED: add other worksheets
    sheet_names =wb.get_sheet_names()
    for sheet_id in range(len(sheet_names)):
        update_count = 0
        keywords_1 = keywords_1s[sheet_id]
        # set default anchor (null)
        if anchor_1 == [[]]:
            anchor_1 = [ [] for i in range(len(keywords_1)) ]
        if anchor_2 == [[]]:
            anchor_2 = [ [] for i in range(len(keywords_2)) ]
        print("*** " + sheet_names[sheet_id] + " ***")
        ws =wb.get_sheet_by_name(sheet_names[sheet_id])
        # ws = wb.active
        row_num = ws.max_row
        col_num = ws.max_column
        row_split = []
        col_split = []
        # data_mat = [ [ 0 for i in range(len(keywords_2)) ] for j in range(len(keywords_1)) ]  # MODIFIED: comment out this line.
        # search for "续表" to divide to sheet into sub-sheets
        for row in range(1, row_num+1):
            for col in range(1, col_num+1):
                cell_value = ws.cell(row=row,column=col).value
                if not (cell_value is None): cell_value = str(cell_value)
                if cell_value:
                    # cell_value = "".join(cell_value.encode("utf8").split())
                    if re.search( r'.*续表.*', cell_value):
                        # "续表" need to be approximately all in the same row or column for this function to work
                        if row>3 and row<row_num: row_split.append(row)
                        if col>3 and col<col_num: col_split.append(col)
                        if row>3 and col>3: print("the location of 续表 is apart from the margins.")
                    if re.search( r'.*单位.*', cell_value):
                        print(cell_value)
                    if re.search( r'.*抽样.*', cell_value):
                        print(cell_value)
        if row_split: row_split = sorted(list(set(row_split)))
        if col_split: col_split = sorted(list(set(col_split)))
        row_split = [1]+row_split+[row_num+1]
        col_split = [1]+col_split+[col_num+1]
        # search for keywords
        for i in range(0,len(row_split)-1):
            for j in range(0,len(col_split)-1):
                # in each sub-sheet
                kw1_coo = [ 0 for k in range(len(keywords_1)) ]     # coodinate(row and column) of each keyword
                kw2_coo = [ 0 for k in range(len(keywords_2)) ]
                # search each cell in a sub-sheet
                for row in range(row_split[i], row_split[i+1]):
                    for col in range(col_split[j], col_split[j+1]):
                        cell_value = ws.cell(row=row,column=col).value
                        if not (cell_value is None): cell_value = str(cell_value)
                        if cell_value:
                            # cell_value = "".join(cell_value.encode("utf8").split()).replace('\xc2\xa0','').replace('\xe3\x80\x80','').replace('\xef\xbc\x83','').replace('\xe2\x80\x83','').replace('_x000D_','').replace('#','')
                            cell_value = re.sub(r'_.*', '', cell_value)
                            # try matching the cell value with each keyword in keywords_1
                            for n in range(len(keywords_1)):
                                for alias in keywords_1[n]:
                                    if re.search( r'^'+alias+'$', cell_value):
                                        # check if it match the anchor condition
                                        if len(anchor_1[n])>0:
                                            if row+anchor_1[n][0]>0 and col+anchor_1[n][1]>0:
                                                anchor_value = ws.cell(row=row+anchor_1[n][0],column=col+anchor_1[n][1]).value
                                            else:
                                                anchor_value = None
                                            if not (anchor_value is None): anchor_value = str(anchor_value)
                                            if anchor_value:
                                                # anchor_value = "".join(anchor_value.encode("utf8").split()).replace('\xc2\xa0','').replace('\xe3\x80\x80','').replace('\xef\xbc\x83','').replace('\xe2\x80\x83','').replace('_x000D_','').replace('#','')
                                                anchor_value = re.sub(r'_.*', '', anchor_value)
                                                if re.search( r'^'+anchor_1[n][2]+'.*', anchor_value):
                                                    insert_list_1D(kw1_coo, n, (row, col))
                                                    break
                                        else:
                                            insert_list_1D(kw1_coo, n, (row, col))
                                            break
                            # try matching the cell value with each keyword in keywords_2 if none in keywords_1 matches
                            else:
                                for n in range(len(keywords_2)):
                                    for alias in keywords_2[n]:
                                        if re.search( r'^'+alias+'$', cell_value):
                                        # check if it match the anchor condition
                                            if len(anchor_2[n])>0:
                                                if row+anchor_2[n][0]>0 and col+anchor_2[n][1]>0:
                                                    anchor_value = ws.cell(row=row+anchor_2[n][0],column=col+anchor_2[n][1]).value
                                                else:
                                                    anchor_value = None
                                                if not (anchor_value is None): anchor_value = str(anchor_value)
                                                if anchor_value:
                                                    # anchor_value = "".join(anchor_value.encode("utf8").split()).replace('\xc2\xa0','').replace('\xe3\x80\x80','').replace('\xef\xbc\x83','').replace('\xe2\x80\x83','').replace('_x000D_','').replace('#','')
                                                    anchor_value = re.sub(r'_.*', '', anchor_value)
                                                    if re.search( r'^'+anchor_2[n][2]+'.*', anchor_value):
                                                        insert_list_1D(kw2_coo, n, (row, col))
                                                        break
                                            else:
                                                insert_list_1D(kw2_coo, n, (row, col))
                                                break
                # update the value in the cross point of keywords_1 and keywords_2
                for n in range(len(keywords_1)):
                    # MODIFIED: add print
                    print("\nupdating " + keywords_1[n][0] + ": ", end = '')
                    for m in range(len(keywords_2)):
                        if kw1_coo[n]!=0 and kw2_coo[m]!=0:     # if both keywords are found in the sheet
                            # update the right-down cross point cell
                            cell_value = ws.cell(row=kw1_coo[n][0],column=kw2_coo[m][1]).value
                            # MODIFIED: add the following
                            if cell_value is None:
                                cell_value = 0
                            else:
                                cell_value = convert_int(cell_value)
                            if isinstance(scores[sheet_id][n], int):
                                ws.cell(row=kw1_coo[n][0],column=kw2_coo[m][1]).value = cell_value + scores[sheet_id][n]
                                update_count = update_count+1
                                print(str(cell_value) + " -> " + str(cell_value + scores[sheet_id][n]), end = ', ')
                            else:
                                print("# WARNING: score undefined.")
                            # MODIFIED: comment out the following
                            '''
                            if not (isinstance(cell_value, unicode) or cell_value is None): cell_value = str(cell_value)
                            if cell_value:
                                cell_value = "".join(cell_value.encode("utf8").split())
                                if re.match( r'^-?\d+\.?\d*$', cell_value):
                                    insert_list_2D(data_mat, n, m, float(cell_value))
                                else:
                                    print (kw1_coo[n][0],kw2_coo[m][1]), cell_value
                                    exit_alert('Error format in data cell!')
                            else:
                                print "*** Empty cell found: ", (kw1_coo[n][0], kw2_coo[m][1])
                            # try the left-up cross point cell
                            cell_value = ws.cell(row=kw2_coo[m][0],column=kw1_coo[n][1]).value
                            if not (isinstance(cell_value, unicode) or cell_value is None): cell_value = str(cell_value)
                            if cell_value:
                                cell_value = "".join(cell_value.encode("utf8").split())
                                if re.match( r'^-?\d+\.?\d*$', cell_value):
                                    insert_list_2D(data_mat, n, m, float(cell_value))
                                    print "*** Left-up cell inserted: ", (kw1_coo[m][0], kw2_coo[n][1])
                            '''
                            # MODIFIED: add the following
                        else:
                            print("# WARNING: not in the gradesheet.")
            print("\n\n" + keywords_2[m][0] + " Finished: " + str(update_count) + " updated in total.\n")
    wb.save(file_path)
    # return data_mat

if __name__ == "__main__":
    # parameters for Math10A
    row_max = 600   #set a upper bound of the row number of the xlsx file by hand due to potential bugs in openpyxl
    file_keywords = ['HW'+str(i) for i in range(36)] + ['Mid1', 'Mid2', 'Final']

    # for exams only, modify this for your sections
    sec1_emails = ['modify this for your sections']  # a list of emails in your first section ("first" mean the first worksheet in the gradesheet), e.g. ['abc@berkeley.edu', 'def@berkeley.edu', ...]. You can get the email list on Calcentral, and then apply a regex replacement
    sec1_sids = ['modify this for your sections']  # a list of corresponding sid in your first section. sids and emails should match in order. Should be easy to get by a regex replacement "\n" -> ", "
    sec2_emails = ['modify this for your sections']  # a list of sid in your second section
    sec2_sids = ['modify this for your sections']  # a list of sid in your second section. sids and emails should match in order.
    section_No = ['modify this for your sections'] # first and second section number, e.g. [209, 210]


    exam = False    # use this when uploading hw grades
    # exam = True     # use this when uploading exam grades
    if exam:
        score_col = 3
        email_col = 2
    else:
        score_col = 2
        sec_col = 6
        sid_col = 7


    gradesheet = './Gradesheet.xlsx'

    # read section names from Gradesheet.xlsx
    wb = load_workbook(filename=gradesheet, read_only=True)
    sheet_names =wb.get_sheet_names()
    sections = []
    for sheet_name in sheet_names:
        ws =wb.get_sheet_by_name(sheet_name)
        sec_name = str(ws.cell(row=1,column=17).value)
        sections.append(int(re.sub("[^0-9]", "", sec_name)))

    # find all data files (HW1.xlsx etc) occurs
    files = []
    file_showup = []
    for file_keyword in file_keywords:
        file_path = glob.glob("./*"+file_keyword+"*xlsx")
        if len(file_path) == 1:
            files.append(file_path[0])
            file_showup.append(file_keyword)

    # read data from each data file, and update the gradesheet
    for file_ind in range(len(files)):
        wb = load_workbook(filename=files[file_ind], read_only=True)
        ws = wb.active
        row_num = min(ws.max_row, row_max)
        col_num = ws.max_column

        scores = [[] for i in sections]
        sids = [[] for i in sections]

        print("\n************** " + file_showup[file_ind] + " **************\n" "\nreading row: ")
        for row in range(1, row_num+1):
            print(str(row)+"       ", end = '\r')
            sys.stdout.flush()
            score = convert_int(ws.cell(row=row,column=score_col).value)
            if exam:
                email = ws.cell(row=row,column=email_col).value
                if email in sec1_emails:
                    section = section_No[0]
                    sid = sec1_sids[sec1_emails.index(email)]
                elif email in sec2_emails:
                    section = section_No[1]
                    sid = sec2_sids[sec2_emails.index(email)]
                else: section = 'no'
            else:
                sid = convert_int(ws.cell(row=row,column=sid_col).value)
                section = convert_int(ws.cell(row=row,column=sec_col).value)
            if section in sections and isinstance(sid, int):
                sec_ind = sections.index(section)
                sids[sec_ind].append(str(sid))
                scores[sec_ind].append(score)

        # update the gradesheet
        print('\n\nUpdating cells in Gradesheet...\n')
        keywords_1s = [form_alias(sid_sec) for sid_sec in sids]
        data_table(file_path=gradesheet, keywords_1s=keywords_1s, keywords_2=[[file_showup[file_ind]]], scores=scores, anchor_1=[[]], anchor_2=[[]]);




















#
